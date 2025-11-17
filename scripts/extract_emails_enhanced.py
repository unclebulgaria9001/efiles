"""
Enhanced Epstein Email PDF Extraction Script
- Extracts and organizes email content with detailed metadata
- Generates content summaries and classifications
- Performs OCR on images
- Creates structured, searchable outputs
"""

import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from collections import Counter
import traceback

# PDF processing libraries
import pdfplumber
import PyPDF2
from PIL import Image
import pytesseract

# Excel and data processing
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('email_extraction_enhanced.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class EnhancedEmailExtractor:
    """Enhanced email extractor with content analysis and OCR"""
    
    def __init__(self, base_dir: str):
        self.base_dir = Path(base_dir)
        self.extracted_dir = self.base_dir / "extracted"
        self.output_file = self.base_dir / f"epstein_emails_enhanced_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.progress_file = self.base_dir / "extraction_progress_enhanced.json"
        self.text_output_dir = self.base_dir / "extracted_emails_organized"
        self.text_output_dir.mkdir(exist_ok=True)
        self.emails_data = []
        self.processed_files = self.load_progress()
        
        # Keywords for content classification
        self.keywords = {
            'legal': ['subpoena', 'deposition', 'court', 'lawsuit', 'attorney', 'counsel', 'motion', 'filing'],
            'media': ['press', 'journalist', 'article', 'publication', 'reporter', 'news', 'times', 'post'],
            'correspondence': ['meeting', 'schedule', 'calendar', 'appointment', 'call', 'discuss'],
            'names': ['maxwell', 'giuffre', 'epstein', 'dershowitz', 'prince', 'andrew'],
            'locations': ['island', 'florida', 'new york', 'london', 'paris', 'virgin islands']
        }
        
    def load_progress(self) -> set:
        """Load previously processed files"""
        if self.progress_file.exists():
            try:
                with open(self.progress_file, 'r') as f:
                    data = json.load(f)
                    return set(data.get('processed_files', []))
            except Exception as e:
                logger.warning(f"Could not load progress file: {e}")
        return set()
    
    def save_progress(self, filename: str):
        """Save progress to resume if interrupted"""
        self.processed_files.add(filename)
        try:
            with open(self.progress_file, 'w') as f:
                json.dump({
                    'processed_files': list(self.processed_files),
                    'last_updated': datetime.now().isoformat(),
                    'total_emails': len(self.emails_data)
                }, f, indent=2)
        except Exception as e:
            logger.error(f"Could not save progress: {e}")
    
    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract text from PDF using multiple methods"""
        text = ""
        
        # Method 1: Try pdfplumber first
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            logger.warning(f"pdfplumber failed for {pdf_path.name}: {e}")
        
        # Method 2: Fallback to PyPDF2
        if not text.strip():
            try:
                with open(pdf_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
            except Exception as e:
                logger.warning(f"PyPDF2 failed for {pdf_path.name}: {e}")
        
        return text.strip()
    
    def extract_images_and_ocr(self, pdf_path: Path) -> Tuple[str, List[str]]:
        """Extract images from PDF and perform OCR"""
        ocr_text = ""
        image_descriptions = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    # Convert page to image for OCR
                    try:
                        img = page.to_image(resolution=300)
                        pil_image = img.original
                        
                        # Perform OCR
                        page_ocr = pytesseract.image_to_string(pil_image)
                        if page_ocr.strip():
                            ocr_text += f"\n--- Page {page_num} OCR ---\n{page_ocr}\n"
                            image_descriptions.append(f"Page {page_num}: {len(page_ocr)} characters extracted")
                    except Exception as e:
                        logger.debug(f"OCR failed for page {page_num} of {pdf_path.name}: {e}")
                        
        except Exception as e:
            logger.warning(f"Image extraction failed for {pdf_path.name}: {e}")
        
        return ocr_text, image_descriptions
    
    def analyze_content(self, text: str) -> Dict[str, any]:
        """Analyze content and generate metadata"""
        text_lower = text.lower()
        
        analysis = {
            'word_count': len(text.split()),
            'char_count': len(text),
            'categories': [],
            'key_people': [],
            'key_locations': [],
            'contains_attachments': 'attachment' in text_lower or 'attached' in text_lower,
            'is_forwarded': 'fwd:' in text_lower or 'forwarded' in text_lower,
            'is_reply': 're:' in text_lower or 'reply' in text_lower,
        }
        
        # Categorize content
        for category, keywords in self.keywords.items():
            if any(keyword in text_lower for keyword in keywords):
                if category == 'names':
                    analysis['key_people'].extend([kw for kw in keywords if kw in text_lower])
                elif category == 'locations':
                    analysis['key_locations'].extend([kw for kw in keywords if kw in text_lower])
                else:
                    analysis['categories'].append(category)
        
        # Remove duplicates
        analysis['key_people'] = list(set(analysis['key_people']))
        analysis['key_locations'] = list(set(analysis['key_locations']))
        
        return analysis
    
    def extract_email_participants(self, text: str) -> Dict[str, List[str]]:
        """Extract email participants (From, To, CC)"""
        participants = {
            'from': [],
            'to': [],
            'cc': []
        }
        
        # Patterns for email participants
        from_pattern = r'From:\s*([^\n]+)'
        to_pattern = r'To:\s*([^\n]+)'
        cc_pattern = r'Cc:\s*([^\n]+)'
        
        from_matches = re.findall(from_pattern, text, re.IGNORECASE)
        to_matches = re.findall(to_pattern, text, re.IGNORECASE)
        cc_matches = re.findall(cc_pattern, text, re.IGNORECASE)
        
        participants['from'] = [m.strip() for m in from_matches if m.strip()]
        participants['to'] = [m.strip() for m in to_matches if m.strip()]
        participants['cc'] = [m.strip() for m in cc_matches if m.strip()]
        
        return participants
    
    def generate_summary(self, email_data: Dict) -> str:
        """Generate a concise summary of the email content"""
        summary_parts = []
        
        # Basic info
        if email_data.get('subject'):
            summary_parts.append(f"Subject: {email_data['subject']}")
        
        if email_data.get('date'):
            summary_parts.append(f"Date: {email_data['date']}")
        
        # Participants
        participants = email_data.get('participants', {})
        if participants.get('from'):
            summary_parts.append(f"From: {', '.join(participants['from'][:3])}")
        if participants.get('to'):
            summary_parts.append(f"To: {', '.join(participants['to'][:3])}")
        
        # Content analysis
        analysis = email_data.get('analysis', {})
        if analysis.get('categories'):
            summary_parts.append(f"Categories: {', '.join(analysis['categories'])}")
        
        if analysis.get('key_people'):
            summary_parts.append(f"Key People: {', '.join(analysis['key_people'][:5])}")
        
        # Thread info
        reply_count = len(email_data.get('replies', []))
        if reply_count > 0:
            summary_parts.append(f"Thread: {reply_count} replies")
        
        # Content type
        content_type = []
        if analysis.get('is_forwarded'):
            content_type.append("Forwarded")
        if analysis.get('is_reply'):
            content_type.append("Reply")
        if analysis.get('contains_attachments'):
            content_type.append("Has Attachments")
        if content_type:
            summary_parts.append(f"Type: {', '.join(content_type)}")
        
        return " | ".join(summary_parts) if summary_parts else "No summary available"
    
    def parse_email_thread(self, text: str) -> Dict[str, any]:
        """Parse email thread into structured data"""
        email_data = {
            'subject': '',
            'date': '',
            'initial_content': '',
            'replies': [],
            'participants': {},
            'analysis': {},
            'summary': ''
        }
        
        # Extract subject
        subject_patterns = [
            r'Subject:\s*(.+?)(?:\n|$)',
            r'Re:\s*(.+?)(?:\n|$)',
            r'Fwd:\s*(.+?)(?:\n|$)',
        ]
        for pattern in subject_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                email_data['subject'] = match.group(1).strip()
                break
        
        # Extract date
        date_patterns = [
            r'Date:\s*(.+?)(?:\n|$)',
            r'Sent:\s*(.+?)(?:\n|$)',
            r'(\w+,\s+\w+\s+\d{1,2},?\s+\d{4})',
            r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        ]
        for pattern in date_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                email_data['date'] = match.group(1).strip()
                break
        
        # Extract participants
        email_data['participants'] = self.extract_email_participants(text)
        
        # Split email thread
        separators = [
            r'-{3,}\s*Original Message\s*-{3,}',
            r'-{3,}\s*Forwarded message\s*-{3,}',
            r'From:\s*.+?\n',
            r'On\s+.+?wrote:',
            r'_{10,}',
            r'={10,}'
        ]
        
        parts = [text]
        for separator in separators:
            new_parts = []
            for part in parts:
                split_parts = re.split(separator, part, flags=re.IGNORECASE)
                new_parts.extend(split_parts)
            parts = new_parts
        
        parts = [p.strip() for p in parts if p.strip()]
        
        if parts:
            email_data['initial_content'] = parts[0][:10000]
            email_data['replies'] = [p[:10000] for p in parts[1:]]
        
        # Analyze content
        email_data['analysis'] = self.analyze_content(text)
        
        # Generate summary
        email_data['summary'] = self.generate_summary(email_data)
        
        return email_data
    
    def process_pdf(self, pdf_path: Path) -> Optional[Dict]:
        """Process a single PDF file with enhanced extraction"""
        try:
            logger.info(f"Processing: {pdf_path.name}")
            
            # Extract text
            text = self.extract_text_from_pdf(pdf_path)
            
            # Try OCR if no text found
            ocr_text = ""
            image_info = []
            if not text.strip():
                logger.info(f"No text found, attempting OCR for {pdf_path.name}")
                ocr_text, image_info = self.extract_images_and_ocr(pdf_path)
                text = ocr_text
            
            if not text.strip():
                logger.warning(f"No text extracted from {pdf_path.name}")
                return None
            
            # Parse email content
            email_data = self.parse_email_thread(text)
            email_data['filename'] = pdf_path.name
            email_data['full_text'] = text[:20000]  # Store more text
            email_data['ocr_performed'] = bool(ocr_text)
            email_data['image_info'] = image_info
            
            return email_data
            
        except Exception as e:
            logger.error(f"Error processing {pdf_path.name}: {e}")
            logger.error(traceback.format_exc())
            return None
    
    def sanitize_filename(self, text: str, max_length: int = 100) -> str:
        """Create a safe filename from text"""
        if not text:
            return "untitled"
        
        text = re.sub(r'[<>:"/\\|?*]', '_', text)
        text = ' '.join(text.split())
        if len(text) > max_length:
            text = text[:max_length]
        text = text.rstrip('. ')
        
        return text if text else "untitled"
    
    def save_organized_text_file(self, email_data: Dict, index: int):
        """Save email to organized text file with comprehensive information"""
        try:
            # Create descriptive filename with From/To/Date/Subject
            subject = email_data.get('subject', 'No_Subject')
            date = email_data.get('date', 'No_Date')
            participants = email_data.get('participants', {})
            
            # Get first From and To
            from_list = participants.get('from', [])
            to_list = participants.get('to', [])
            
            # Extract just the name (before email address if present)
            from_name = 'Unknown'
            if from_list:
                from_text = from_list[0]
                # Try to extract name before email
                name_match = re.match(r'^([^<@]+)', from_text)
                if name_match:
                    from_name = name_match.group(1).strip()
                else:
                    from_name = from_text[:30]
            
            to_name = 'Unknown'
            if to_list:
                to_text = to_list[0]
                # Try to extract name before email
                name_match = re.match(r'^([^<@]+)', to_text)
                if name_match:
                    to_name = name_match.group(1).strip()
                else:
                    to_name = to_text[:30]
            
            # Sanitize components
            safe_from = self.sanitize_filename(from_name, max_length=25)
            safe_to = self.sanitize_filename(to_name, max_length=25)
            safe_date = self.sanitize_filename(date, max_length=20)
            safe_subject = self.sanitize_filename(subject, max_length=40)
            
            # Create filename: index_date_from_to_subject.txt
            filename = f"{index:04d}_{safe_date}_From-{safe_from}_To-{safe_to}_{safe_subject}.txt"
            filepath = self.text_output_dir / filename
            
            # Prepare organized content
            lines = []
            lines.append("=" * 100)
            lines.append(f"EMAIL #{index:04d} - ORGANIZED EXTRACTION")
            lines.append("=" * 100)
            lines.append("")
            
            # METADATA SECTION
            lines.append("╔" + "═" * 98 + "╗")
            lines.append("║" + " METADATA ".center(98) + "║")
            lines.append("╚" + "═" * 98 + "╝")
            lines.append("")
            lines.append(f"Source PDF:        {email_data.get('filename', 'N/A')}")
            lines.append(f"Subject:           {email_data.get('subject', 'N/A')}")
            lines.append(f"Date:              {email_data.get('date', 'N/A')}")
            lines.append(f"OCR Performed:     {'Yes' if email_data.get('ocr_performed') else 'No'}")
            lines.append("")
            
            # PARTICIPANTS SECTION
            participants = email_data.get('participants', {})
            if any(participants.values()):
                lines.append("╔" + "═" * 98 + "╗")
                lines.append("║" + " PARTICIPANTS ".center(98) + "║")
                lines.append("╚" + "═" * 98 + "╝")
                lines.append("")
                if participants.get('from'):
                    lines.append(f"From:  {', '.join(participants['from'])}")
                if participants.get('to'):
                    lines.append(f"To:    {', '.join(participants['to'])}")
                if participants.get('cc'):
                    lines.append(f"CC:    {', '.join(participants['cc'])}")
                lines.append("")
            
            # CONTENT ANALYSIS SECTION
            analysis = email_data.get('analysis', {})
            lines.append("╔" + "═" * 98 + "╗")
            lines.append("║" + " CONTENT ANALYSIS ".center(98) + "║")
            lines.append("╚" + "═" * 98 + "╝")
            lines.append("")
            lines.append(f"Word Count:        {analysis.get('word_count', 0):,}")
            lines.append(f"Character Count:   {analysis.get('char_count', 0):,}")
            lines.append(f"Categories:        {', '.join(analysis.get('categories', [])) or 'None identified'}")
            lines.append(f"Key People:        {', '.join(analysis.get('key_people', [])) or 'None identified'}")
            lines.append(f"Key Locations:     {', '.join(analysis.get('key_locations', [])) or 'None identified'}")
            lines.append(f"Type:              {'Reply' if analysis.get('is_reply') else ''} {'Forwarded' if analysis.get('is_forwarded') else ''} {'Has Attachments' if analysis.get('contains_attachments') else ''}".strip() or 'Standard')
            lines.append(f"Thread Replies:    {len(email_data.get('replies', []))}")
            lines.append("")
            
            # SUMMARY SECTION
            lines.append("╔" + "═" * 98 + "╗")
            lines.append("║" + " SUMMARY ".center(98) + "║")
            lines.append("╚" + "═" * 98 + "╝")
            lines.append("")
            lines.append(email_data.get('summary', 'No summary available'))
            lines.append("")
            lines.append("")
            
            # INITIAL EMAIL CONTENT
            lines.append("╔" + "═" * 98 + "╗")
            lines.append("║" + " INITIAL EMAIL CONTENT ".center(98) + "║")
            lines.append("╚" + "═" * 98 + "╝")
            lines.append("")
            lines.append(email_data.get('initial_content', 'N/A'))
            lines.append("")
            
            # REPLIES
            replies = email_data.get('replies', [])
            if replies:
                for i, reply in enumerate(replies, 1):
                    lines.append("")
                    lines.append("╔" + "═" * 98 + "╗")
                    lines.append("║" + f" REPLY #{i} ".center(98) + "║")
                    lines.append("╚" + "═" * 98 + "╝")
                    lines.append("")
                    lines.append(reply)
                    lines.append("")
            
            # IMAGE/OCR INFO
            if email_data.get('image_info'):
                lines.append("")
                lines.append("╔" + "═" * 98 + "╗")
                lines.append("║" + " OCR / IMAGE INFORMATION ".center(98) + "║")
                lines.append("╚" + "═" * 98 + "╝")
                lines.append("")
                for info in email_data['image_info']:
                    lines.append(f"• {info}")
                lines.append("")
            
            # FULL TEXT
            lines.append("")
            lines.append("╔" + "═" * 98 + "╗")
            lines.append("║" + " COMPLETE EXTRACTED TEXT ".center(98) + "║")
            lines.append("╚" + "═" * 98 + "╝")
            lines.append("")
            lines.append(email_data.get('full_text', 'N/A'))
            
            # Write to file
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            logger.info(f"Saved organized file: {filename}")
            
        except Exception as e:
            logger.error(f"Error saving text file: {e}")
    
    def process_all_pdfs(self):
        """Process all PDF files"""
        pdf_files = list(self.extracted_dir.rglob("*.pdf"))
        total_files = len(pdf_files)
        
        logger.info(f"Found {total_files} PDF files to process")
        
        for idx, pdf_path in enumerate(pdf_files, 1):
            if str(pdf_path) in self.processed_files:
                logger.info(f"Skipping already processed file: {pdf_path.name}")
                continue
            
            logger.info(f"Processing {idx}/{total_files}: {pdf_path.name}")
            
            email_data = self.process_pdf(pdf_path)
            if email_data:
                self.emails_data.append(email_data)
                self.save_organized_text_file(email_data, idx)
            
            if idx % 10 == 0:
                self.save_progress(str(pdf_path))
                self.save_to_excel()
                logger.info(f"Progress saved: {idx}/{total_files} files processed")
        
        logger.info(f"Completed processing all {total_files} PDF files")
    
    def save_to_excel(self):
        """Save to Excel with enhanced organization"""
        if not self.emails_data:
            logger.warning("No email data to save")
            return
        
        logger.info(f"Saving {len(self.emails_data)} emails to Excel...")
        
        rows = []
        max_replies = max(len(email.get('replies', [])) for email in self.emails_data)
        
        for email in self.emails_data:
            analysis = email.get('analysis', {})
            participants = email.get('participants', {})
            
            row = {
                'Filename': email.get('filename', ''),
                'Summary': email.get('summary', ''),
                'Subject': email.get('subject', ''),
                'Date': email.get('date', ''),
                'From': '; '.join(participants.get('from', [])),
                'To': '; '.join(participants.get('to', [])),
                'CC': '; '.join(participants.get('cc', [])),
                'Categories': ', '.join(analysis.get('categories', [])),
                'Key People': ', '.join(analysis.get('key_people', [])),
                'Key Locations': ', '.join(analysis.get('key_locations', [])),
                'Word Count': analysis.get('word_count', 0),
                'Reply Count': len(email.get('replies', [])),
                'Type': ('Reply ' if analysis.get('is_reply') else '') + ('Forwarded ' if analysis.get('is_forwarded') else '') + ('Attachments' if analysis.get('contains_attachments') else ''),
                'OCR Used': 'Yes' if email.get('ocr_performed') else 'No',
                'Initial Content': email.get('initial_content', ''),
            }
            
            # Add reply columns
            replies = email.get('replies', [])
            for i in range(max_replies):
                row[f'Reply {i+1}'] = replies[i] if i < len(replies) else ''
            
            row['Full Text'] = email.get('full_text', '')
            
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Emails', index=False)
            
            worksheet = writer.sheets['Emails']
            
            # Format headers
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Adjust column widths
            column_widths = {
                'Filename': 35,
                'Summary': 60,
                'Subject': 40,
                'Date': 25,
                'From': 35,
                'To': 35,
                'CC': 30,
                'Categories': 25,
                'Key People': 30,
                'Key Locations': 25,
                'Word Count': 12,
                'Reply Count': 12,
                'Type': 20,
                'OCR Used': 10,
                'Initial Content': 50,
                'Full Text': 50
            }
            
            for idx, col in enumerate(df.columns, 1):
                column_letter = get_column_letter(idx)
                width = column_widths.get(col, 40)
                worksheet.column_dimensions[column_letter].width = width
            
            # Enable text wrapping
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        logger.info(f"Excel file saved: {self.output_file}")
        logger.info(f"Total emails: {len(self.emails_data)}")
    
    def run(self):
        """Main execution"""
        logger.info("=" * 80)
        logger.info("Starting Enhanced Email Extraction Process")
        logger.info("=" * 80)
        logger.info(f"Base directory: {self.base_dir}")
        logger.info(f"Output file: {self.output_file}")
        logger.info(f"Text files: {self.text_output_dir}")
        
        start_time = datetime.now()
        
        try:
            self.process_all_pdfs()
            self.save_to_excel()
            
            end_time = datetime.now()
            duration = end_time - start_time
            
            logger.info("=" * 80)
            logger.info("Enhanced Extraction Complete!")
            logger.info(f"Total time: {duration}")
            logger.info(f"Total emails: {len(self.emails_data)}")
            logger.info(f"Excel file: {self.output_file}")
            logger.info(f"Text files: {self.text_output_dir}")
            logger.info("=" * 80)
            
        except Exception as e:
            logger.error(f"Fatal error: {e}")
            logger.error(traceback.format_exc())
            raise


def main():
    """Main entry point"""
    base_dir = r"G:\My Drive\04_Resources\Notes\Epstein Email Dump"
    
    extractor = EnhancedEmailExtractor(base_dir)
    extractor.run()


if __name__ == "__main__":
    main()

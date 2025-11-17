"""
Coded Message Analysis Script
Analyzes emails for potential coded messages, hidden communications, and suspicious patterns
"""

import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from collections import Counter, defaultdict
import pandas as pd

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('coded_message_analysis.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class CodedMessageAnalyzer:
    """Analyzes emails for coded messages and suspicious patterns"""
    
    def __init__(self, base_dir: str):
        self.base_dir = Path(base_dir)
        self.text_dir = self.base_dir / "extracted_emails_organized"
        self.output_file = self.base_dir / f"coded_messages_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.findings = []
        
        # Suspicious patterns and code indicators
        self.code_indicators = {
            'unusual_capitalization': [],
            'repeated_phrases': [],
            'unusual_punctuation': [],
            'number_sequences': [],
            'time_references': [],
            'location_codes': [],
            'unusual_words': [],
            'acronyms': [],
            'suspicious_phrases': []
        }
        
        # Known code words and suspicious terminology
        self.suspicious_terms = {
            'food_codes': ['pizza', 'hotdog', 'pasta', 'sauce', 'cheese', 'walnut', 'map'],
            'time_codes': ['late night', 'early morning', 'midnight', 'dawn'],
            'vague_references': ['package', 'delivery', 'shipment', 'cargo', 'item', 'gift'],
            'euphemisms': ['entertainment', 'party', 'special guest', 'friend', 'massage'],
            'locations': ['island', 'ranch', 'estate', 'private', 'secluded', 'remote'],
            'age_references': ['young', 'fresh', 'new', 'teen', 'minor', 'underage'],
            'secrecy': ['confidential', 'private', 'discreet', 'quiet', 'secret', 'hidden']
        }
        
        # Unusual patterns
        self.pattern_checks = {
            'excessive_deletions': r'\[REDACTED\]|\[DELETED\]|\*\*\*',
            'unusual_spacing': r'\s{3,}',
            'hidden_text': r'\[.*?\]|\(.*?\)',
            'number_codes': r'\b\d{2,}\b',
            'time_stamps': r'\d{1,2}:\d{2}\s*(?:AM|PM|am|pm)?',
            'coordinates': r'\d+\.\d+\s*[NS]\s*,\s*\d+\.\d+\s*[EW]',
            'phone_numbers': r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b',
            'unusual_emails': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        }
        
    def analyze_capitalization(self, text: str) -> List[Dict]:
        """Detect unusual capitalization patterns that might be codes"""
        findings = []
        
        # Find words with unusual capitalization
        words = re.findall(r'\b[A-Z][a-z]*[A-Z][a-z]*\b', text)
        if words:
            findings.append({
                'type': 'Unusual Capitalization',
                'pattern': 'Mixed case words',
                'examples': list(set(words))[:10],
                'count': len(words),
                'suspicion_level': 'Medium' if len(words) > 5 else 'Low'
            })
        
        # Find all-caps words (potential emphasis or code)
        all_caps = re.findall(r'\b[A-Z]{3,}\b', text)
        if len(all_caps) > 3:
            findings.append({
                'type': 'All Caps Words',
                'pattern': 'Excessive capitalization',
                'examples': list(set(all_caps))[:10],
                'count': len(all_caps),
                'suspicion_level': 'Low'
            })
        
        # Find first letters of sentences (potential acrostic)
        sentences = re.split(r'[.!?]\s+', text)
        if len(sentences) > 5:
            first_letters = ''.join([s[0] for s in sentences if s])
            if len(first_letters) > 10:
                findings.append({
                    'type': 'Potential Acrostic',
                    'pattern': 'First letters of sentences',
                    'examples': [first_letters[:50]],
                    'count': len(first_letters),
                    'suspicion_level': 'Low'
                })
        
        return findings
    
    def analyze_suspicious_terms(self, text: str) -> List[Dict]:
        """Detect suspicious terminology and potential code words"""
        findings = []
        text_lower = text.lower()
        
        for category, terms in self.suspicious_terms.items():
            found_terms = [term for term in terms if term in text_lower]
            if found_terms:
                # Count occurrences
                term_counts = {term: text_lower.count(term) for term in found_terms}
                total_count = sum(term_counts.values())
                
                suspicion = 'High' if total_count > 5 else 'Medium' if total_count > 2 else 'Low'
                
                findings.append({
                    'type': f'Suspicious Terms: {category}',
                    'pattern': f'Known code word category',
                    'examples': [f"{term} ({count}x)" for term, count in term_counts.items()],
                    'count': total_count,
                    'suspicion_level': suspicion
                })
        
        return findings
    
    def analyze_number_patterns(self, text: str) -> List[Dict]:
        """Detect unusual number sequences and patterns"""
        findings = []
        
        # Find all number sequences
        numbers = re.findall(r'\b\d+\b', text)
        if len(numbers) > 10:
            # Check for repeated numbers
            number_counts = Counter(numbers)
            repeated = {num: count for num, count in number_counts.items() if count > 2}
            
            if repeated:
                findings.append({
                    'type': 'Repeated Numbers',
                    'pattern': 'Same numbers appearing multiple times',
                    'examples': [f"{num} ({count}x)" for num, count in list(repeated.items())[:10]],
                    'count': len(repeated),
                    'suspicion_level': 'Medium'
                })
        
        # Find time stamps
        times = re.findall(r'\b\d{1,2}:\d{2}\s*(?:AM|PM|am|pm)?\b', text)
        if len(times) > 3:
            findings.append({
                'type': 'Multiple Time References',
                'pattern': 'Unusual number of time stamps',
                'examples': list(set(times))[:10],
                'count': len(times),
                'suspicion_level': 'Low'
            })
        
        # Find phone numbers
        phones = re.findall(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b', text)
        if phones:
            findings.append({
                'type': 'Phone Numbers',
                'pattern': 'Contact information',
                'examples': list(set(phones))[:5],
                'count': len(phones),
                'suspicion_level': 'Low'
            })
        
        return findings
    
    def analyze_repeated_phrases(self, text: str) -> List[Dict]:
        """Detect repeated phrases that might be code signals"""
        findings = []
        
        # Find repeated 2-3 word phrases
        words = text.lower().split()
        bigrams = [' '.join(words[i:i+2]) for i in range(len(words)-1)]
        trigrams = [' '.join(words[i:i+3]) for i in range(len(words)-2)]
        
        bigram_counts = Counter(bigrams)
        trigram_counts = Counter(trigrams)
        
        # Find phrases repeated more than twice
        repeated_bigrams = {phrase: count for phrase, count in bigram_counts.items() if count > 2}
        repeated_trigrams = {phrase: count for phrase, count in trigram_counts.items() if count > 2}
        
        if repeated_bigrams:
            findings.append({
                'type': 'Repeated 2-Word Phrases',
                'pattern': 'Phrases appearing multiple times',
                'examples': [f"{phrase} ({count}x)" for phrase, count in list(repeated_bigrams.items())[:10]],
                'count': len(repeated_bigrams),
                'suspicion_level': 'Medium' if len(repeated_bigrams) > 5 else 'Low'
            })
        
        if repeated_trigrams:
            findings.append({
                'type': 'Repeated 3-Word Phrases',
                'pattern': 'Longer phrases appearing multiple times',
                'examples': [f"{phrase} ({count}x)" for phrase, count in list(repeated_trigrams.items())[:10]],
                'count': len(repeated_trigrams),
                'suspicion_level': 'Medium'
            })
        
        return findings
    
    def analyze_unusual_punctuation(self, text: str) -> List[Dict]:
        """Detect unusual punctuation patterns"""
        findings = []
        
        # Excessive punctuation
        excessive_dots = re.findall(r'\.{3,}', text)
        if len(excessive_dots) > 3:
            findings.append({
                'type': 'Excessive Ellipsis',
                'pattern': 'Multiple dots/ellipsis',
                'examples': excessive_dots[:10],
                'count': len(excessive_dots),
                'suspicion_level': 'Low'
            })
        
        # Unusual spacing
        unusual_spaces = re.findall(r'\s{3,}', text)
        if len(unusual_spaces) > 5:
            findings.append({
                'type': 'Unusual Spacing',
                'pattern': 'Excessive spaces (potential hidden text)',
                'examples': [f"{len(s)} spaces" for s in unusual_spaces[:10]],
                'count': len(unusual_spaces),
                'suspicion_level': 'Medium'
            })
        
        # Brackets and parentheses (potential hidden messages)
        bracketed = re.findall(r'\[([^\]]+)\]', text)
        if len(bracketed) > 5:
            findings.append({
                'type': 'Bracketed Text',
                'pattern': 'Text in brackets (potential codes)',
                'examples': bracketed[:10],
                'count': len(bracketed),
                'suspicion_level': 'Medium'
            })
        
        return findings
    
    def analyze_redactions(self, text: str) -> List[Dict]:
        """Detect redacted or deleted content"""
        findings = []
        
        redaction_patterns = [
            r'\[REDACTED\]',
            r'\[DELETED\]',
            r'\*\*\*+',
            r'###',
            r'\[.*?REMOVED.*?\]'
        ]
        
        total_redactions = 0
        redaction_types = []
        
        for pattern in redaction_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                total_redactions += len(matches)
                redaction_types.extend(matches)
        
        if total_redactions > 0:
            findings.append({
                'type': 'Redacted Content',
                'pattern': 'Removed or censored information',
                'examples': list(set(redaction_types))[:10],
                'count': total_redactions,
                'suspicion_level': 'High' if total_redactions > 5 else 'Medium'
            })
        
        return findings
    
    def analyze_vague_language(self, text: str) -> List[Dict]:
        """Detect vague or evasive language"""
        findings = []
        text_lower = text.lower()
        
        vague_terms = [
            'you know what i mean', 'if you know what i mean', 'wink wink',
            'you understand', 'as discussed', 'as we talked about',
            'the usual', 'same as before', 'like last time',
            'the thing', 'that matter', 'our arrangement',
            'as agreed', 'per our conversation', 'you know'
        ]
        
        found_vague = [term for term in vague_terms if term in text_lower]
        
        if found_vague:
            findings.append({
                'type': 'Vague Language',
                'pattern': 'Evasive or unclear references',
                'examples': found_vague,
                'count': len(found_vague),
                'suspicion_level': 'High' if len(found_vague) > 3 else 'Medium'
            })
        
        return findings
    
    def analyze_email_file(self, filepath: Path) -> Optional[Dict]:
        """Analyze a single email file for coded messages"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Extract email content (skip metadata sections)
            email_content = content
            if 'INITIAL EMAIL CONTENT' in content:
                parts = content.split('INITIAL EMAIL CONTENT')
                if len(parts) > 1:
                    email_content = parts[1]
            
            # Run all analyses
            all_findings = []
            all_findings.extend(self.analyze_capitalization(email_content))
            all_findings.extend(self.analyze_suspicious_terms(email_content))
            all_findings.extend(self.analyze_number_patterns(email_content))
            all_findings.extend(self.analyze_repeated_phrases(email_content))
            all_findings.extend(self.analyze_unusual_punctuation(email_content))
            all_findings.extend(self.analyze_redactions(email_content))
            all_findings.extend(self.analyze_vague_language(email_content))
            
            if all_findings:
                # Calculate overall suspicion score
                suspicion_scores = {'High': 3, 'Medium': 2, 'Low': 1}
                total_score = sum(suspicion_scores.get(f['suspicion_level'], 0) for f in all_findings)
                avg_score = total_score / len(all_findings) if all_findings else 0
                
                overall_suspicion = 'High' if avg_score > 2 else 'Medium' if avg_score > 1.5 else 'Low'
                
                return {
                    'filename': filepath.name,
                    'total_indicators': len(all_findings),
                    'overall_suspicion': overall_suspicion,
                    'suspicion_score': round(avg_score, 2),
                    'findings': all_findings,
                    'high_priority': overall_suspicion == 'High'
                }
            
            return None
            
        except Exception as e:
            logger.error(f"Error analyzing {filepath.name}: {e}")
            return None
    
    def analyze_all_emails(self):
        """Analyze all email files"""
        email_files = list(self.text_dir.glob("*.txt"))
        total_files = len(email_files)
        
        logger.info(f"Analyzing {total_files} email files for coded messages...")
        
        for idx, filepath in enumerate(email_files, 1):
            if idx % 20 == 0:
                logger.info(f"Progress: {idx}/{total_files} files analyzed")
            
            result = self.analyze_email_file(filepath)
            if result:
                self.findings.append(result)
        
        logger.info(f"Analysis complete. Found {len(self.findings)} emails with potential coded messages")
    
    def generate_report(self):
        """Generate comprehensive report"""
        if not self.findings:
            logger.warning("No coded messages found")
            return
        
        # Sort by suspicion level
        self.findings.sort(key=lambda x: (x['suspicion_score'], x['total_indicators']), reverse=True)
        
        # Create detailed report
        rows = []
        for finding in self.findings:
            # Create summary of findings
            finding_types = [f['type'] for f in finding['findings']]
            finding_summary = '; '.join(set(finding_types))
            
            # Get top examples
            examples = []
            for f in finding['findings'][:5]:  # Top 5 findings
                if f['examples']:
                    examples.append(f"{f['type']}: {', '.join(str(e) for e in f['examples'][:3])}")
            
            row = {
                'Filename': finding['filename'],
                'Overall Suspicion': finding['overall_suspicion'],
                'Suspicion Score': finding['suspicion_score'],
                'Total Indicators': finding['total_indicators'],
                'Finding Types': finding_summary,
                'Top Examples': ' | '.join(examples[:3]),
                'High Priority': 'YES' if finding['high_priority'] else 'NO'
            }
            rows.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Coded Messages Analysis', index=False)
            
            worksheet = writer.sheets['Coded Messages Analysis']
            
            # Format headers
            from openpyxl.styles import Font, Alignment, PatternFill
            header_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 80
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 50
            worksheet.column_dimensions['F'].width = 80
            worksheet.column_dimensions['G'].width = 12
            
            # Enable text wrapping
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        logger.info(f"Report saved: {self.output_file}")
        
        # Print summary
        high_priority = sum(1 for f in self.findings if f['high_priority'])
        logger.info(f"\n{'='*80}")
        logger.info(f"CODED MESSAGE ANALYSIS SUMMARY")
        logger.info(f"{'='*80}")
        logger.info(f"Total emails analyzed: {len(list(self.text_dir.glob('*.txt')))}")
        logger.info(f"Emails with potential codes: {len(self.findings)}")
        logger.info(f"High priority emails: {high_priority}")
        logger.info(f"Report file: {self.output_file}")
        logger.info(f"{'='*80}\n")
    
    def run(self):
        """Main execution"""
        logger.info("="*80)
        logger.info("Starting Coded Message Analysis")
        logger.info("="*80)
        
        start_time = datetime.now()
        
        try:
            self.analyze_all_emails()
            self.generate_report()
            
            end_time = datetime.now()
            duration = end_time - start_time
            
            logger.info(f"\nAnalysis completed in {duration}")
            
        except Exception as e:
            logger.error(f"Fatal error: {e}")
            raise


def main():
    """Main entry point"""
    base_dir = r"G:\My Drive\04_Resources\Notes\Epstein Email Dump"
    
    analyzer = CodedMessageAnalyzer(base_dir)
    analyzer.run()


if __name__ == "__main__":
    main()

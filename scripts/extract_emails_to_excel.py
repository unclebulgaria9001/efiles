"""
Epstein Email PDF Extraction Script
Extracts email content from PDFs and organizes into Excel spreadsheet
with columns: Subject, Date, Initial Content, Reply 1, Reply 2, etc.
"""

import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import traceback

# PDF processing libraries
import pdfplumber
import PyPDF2
from PIL import Image
import pytesseract

# Excel and data processing
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('email_extraction.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class EmailExtractor:
    """Extracts and parses email content from PDF files"""
    
    def __init__(self, base_dir: str):
        self.base_dir = Path(base_dir)
        self.extracted_dir = self.base_dir / "extracted"
        self.output_file = self.base_dir / f"epstein_emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.progress_file = self.base_dir / "extraction_progress.json"
        self.text_output_dir = self.base_dir / "extracted_text_files"
        self.text_output_dir.mkdir(exist_ok=True)
        self.emails_data = []
        self.processed_files = self.load_progress()
        
    def load_progress(self) -> set:
        """Load previously processed files"""
        if self.progress_file.exists():
            try:
                with open(self.progress_file, 'r') as f:
                    data = json.load(f)
                    return set(data.get('processed_files', []))
            except Exception as e:
                logger.warning(f"Could not load progress file: {e}")
        return set()
    
    def save_progress(self, filename: str):
        """Save progress to resume if interrupted"""
        self.processed_files.add(filename)
        try:
            with open(self.progress_file, 'w') as f:
                json.dump({
                    'processed_files': list(self.processed_files),
                    'last_updated': datetime.now().isoformat(),
                    'total_emails': len(self.emails_data)
                }, f, indent=2)
        except Exception as e:
            logger.error(f"Could not save progress: {e}")
    
    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract text from PDF using multiple methods"""
        text = ""
        
        # Method 1: Try pdfplumber first (better for formatted text)
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            logger.warning(f"pdfplumber failed for {pdf_path.name}: {e}")
        
        # Method 2: Fallback to PyPDF2
        if not text.strip():
            try:
                with open(pdf_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
            except Exception as e:
                logger.warning(f"PyPDF2 failed for {pdf_path.name}: {e}")
        
        return text.strip()
    
    def extract_images_from_pdf(self, pdf_path: Path) -> List[Image.Image]:
        """Extract images from PDF for OCR"""
        images = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    # Convert page to image
                    img = page.to_image(resolution=300)
                    images.append(img.original)
        except Exception as e:
            logger.warning(f"Image extraction failed for {pdf_path.name}: {e}")
        return images
    
    def ocr_image(self, image: Image.Image) -> str:
        """Perform OCR on image"""
        try:
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            logger.warning(f"OCR failed: {e}")
            return ""
    
    def parse_email_thread(self, text: str) -> Dict[str, any]:
        """Parse email thread into structured data"""
        email_data = {
            'subject': '',
            'date': '',
            'initial_content': '',
            'replies': []
        }
        
        # Common email patterns
        subject_patterns = [
            r'Subject:\s*(.+?)(?:\n|$)',
            r'Re:\s*(.+?)(?:\n|$)',
            r'Fwd:\s*(.+?)(?:\n|$)',
            r'SUBJECT:\s*(.+?)(?:\n|$)'
        ]
        
        date_patterns = [
            r'Date:\s*(.+?)(?:\n|$)',
            r'Sent:\s*(.+?)(?:\n|$)',
            r'DATE:\s*(.+?)(?:\n|$)',
            r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
            r'(\w+\s+\d{1,2},?\s+\d{4})',
            r'(\d{4}-\d{2}-\d{2})'
        ]
        
        # Extract subject
        for pattern in subject_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                email_data['subject'] = match.group(1).strip()
                break
        
        # Extract date
        for pattern in date_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                email_data['date'] = match.group(1).strip()
                break
        
        # Split email thread by common separators
        separators = [
            r'-{3,}\s*Original Message\s*-{3,}',
            r'-{3,}\s*Forwarded message\s*-{3,}',
            r'From:\s*.+?\n',
            r'On\s+.+?wrote:',
            r'_{10,}',
            r'={10,}'
        ]
        
        # Try to split into email parts
        parts = [text]
        for separator in separators:
            new_parts = []
            for part in parts:
                split_parts = re.split(separator, part, flags=re.IGNORECASE)
                new_parts.extend(split_parts)
            parts = new_parts
        
        # Clean and assign parts
        parts = [p.strip() for p in parts if p.strip()]
        
        if parts:
            email_data['initial_content'] = parts[0][:5000]  # Limit to 5000 chars
            email_data['replies'] = [p[:5000] for p in parts[1:]]  # Subsequent replies
        
        return email_data
    
    def process_pdf(self, pdf_path: Path) -> Optional[Dict]:
        """Process a single PDF file"""
        try:
            logger.info(f"Processing: {pdf_path.name}")
            
            # Extract text
            text = self.extract_text_from_pdf(pdf_path)
            
            # If no text found, try OCR
            if not text.strip():
                logger.info(f"No text found, attempting OCR for {pdf_path.name}")
                images = self.extract_images_from_pdf(pdf_path)
                ocr_texts = []
                for img in images[:5]:  # Limit to first 5 pages for OCR
                    ocr_text = self.ocr_image(img)
                    if ocr_text:
                        ocr_texts.append(ocr_text)
                text = "\n".join(ocr_texts)
            
            if not text.strip():
                logger.warning(f"No text extracted from {pdf_path.name}")
                return None
            
            # Parse email content
            email_data = self.parse_email_thread(text)
            email_data['filename'] = pdf_path.name
            email_data['full_text'] = text[:10000]  # Store first 10k chars of full text
            
            return email_data
            
        except Exception as e:
            logger.error(f"Error processing {pdf_path.name}: {e}")
            logger.error(traceback.format_exc())
            return None
    
    def sanitize_filename(self, text: str, max_length: int = 100) -> str:
        """Create a safe filename from text"""
        if not text:
            return "untitled"
        
        # Remove or replace invalid characters
        text = re.sub(r'[<>:"/\\|?*]', '_', text)
        # Remove extra whitespace
        text = ' '.join(text.split())
        # Limit length
        if len(text) > max_length:
            text = text[:max_length]
        # Remove trailing dots and spaces
        text = text.rstrip('. ')
        
        return text if text else "untitled"
    
    def save_text_file(self, email_data: Dict, pdf_filename: str, index: int):
        """Save email content to individual text file with descriptive name"""
        try:
            # Create descriptive filename
            subject = email_data.get('subject', 'No_Subject')
            date = email_data.get('date', 'No_Date')
            
            # Sanitize components
            safe_subject = self.sanitize_filename(subject, max_length=60)
            safe_date = self.sanitize_filename(date, max_length=20)
            
            # Create filename: index_date_subject.txt
            filename = f"{index:04d}_{safe_date}_{safe_subject}.txt"
            filepath = self.text_output_dir / filename
            
            # Prepare content
            content_lines = []
            content_lines.append("=" * 80)
            content_lines.append(f"EMAIL EXTRACTION FROM: {pdf_filename}")
            content_lines.append("=" * 80)
            content_lines.append("")
            content_lines.append(f"Subject: {email_data.get('subject', 'N/A')}")
            content_lines.append(f"Date: {email_data.get('date', 'N/A')}")
            content_lines.append(f"Source PDF: {pdf_filename}")
            content_lines.append("")
            content_lines.append("=" * 80)
            content_lines.append("INITIAL EMAIL CONTENT")
            content_lines.append("=" * 80)
            content_lines.append("")
            content_lines.append(email_data.get('initial_content', 'N/A'))
            content_lines.append("")
            
            # Add replies
            replies = email_data.get('replies', [])
            if replies:
                for i, reply in enumerate(replies, 1):
                    content_lines.append("")
                    content_lines.append("=" * 80)
                    content_lines.append(f"REPLY {i}")
                    content_lines.append("=" * 80)
                    content_lines.append("")
                    content_lines.append(reply)
                    content_lines.append("")
            
            # Add full text section
            content_lines.append("")
            content_lines.append("=" * 80)
            content_lines.append("FULL EXTRACTED TEXT")
            content_lines.append("=" * 80)
            content_lines.append("")
            content_lines.append(email_data.get('full_text', 'N/A'))
            
            # Write to file
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(content_lines))
            
            logger.info(f"Saved text file: {filename}")
            
        except Exception as e:
            logger.error(f"Error saving text file for {pdf_filename}: {e}")
    
    def process_all_pdfs(self):
        """Process all PDF files in the extracted directory"""
        pdf_files = list(self.extracted_dir.rglob("*.pdf"))
        total_files = len(pdf_files)
        
        logger.info(f"Found {total_files} PDF files to process")
        
        for idx, pdf_path in enumerate(pdf_files, 1):
            # Skip if already processed
            if str(pdf_path) in self.processed_files:
                logger.info(f"Skipping already processed file: {pdf_path.name}")
                continue
            
            logger.info(f"Processing {idx}/{total_files}: {pdf_path.name}")
            
            email_data = self.process_pdf(pdf_path)
            if email_data:
                self.emails_data.append(email_data)
                # Save individual text file with descriptive name
                self.save_text_file(email_data, pdf_path.name, idx)
            
            # Save progress every 10 files
            if idx % 10 == 0:
                self.save_progress(str(pdf_path))
                self.save_to_excel()  # Save intermediate results
                logger.info(f"Progress saved: {idx}/{total_files} files processed")
        
        logger.info(f"Completed processing all {total_files} PDF files")
    
    def save_to_excel(self):
        """Save extracted emails to Excel with proper formatting"""
        if not self.emails_data:
            logger.warning("No email data to save")
            return
        
        logger.info(f"Saving {len(self.emails_data)} emails to Excel...")
        
        # Prepare data for DataFrame
        rows = []
        max_replies = max(len(email.get('replies', [])) for email in self.emails_data)
        
        for email in self.emails_data:
            row = {
                'Filename': email.get('filename', ''),
                'Subject': email.get('subject', ''),
                'Date': email.get('date', ''),
                'Initial Content': email.get('initial_content', ''),
            }
            
            # Add reply columns
            replies = email.get('replies', [])
            for i in range(max_replies):
                reply_num = i + 1
                row[f'Reply {reply_num}'] = replies[i] if i < len(replies) else ''
            
            # Add full text column
            row['Full Text'] = email.get('full_text', '')
            
            rows.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Emails', index=False)
            
            # Get the worksheet
            worksheet = writer.sheets['Emails']
            
            # Format headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Adjust column widths
            for idx, col in enumerate(df.columns, 1):
                column_letter = get_column_letter(idx)
                if col == 'Filename':
                    worksheet.column_dimensions[column_letter].width = 30
                elif col in ['Subject', 'Date']:
                    worksheet.column_dimensions[column_letter].width = 25
                elif col == 'Full Text':
                    worksheet.column_dimensions[column_letter].width = 50
                else:
                    worksheet.column_dimensions[column_letter].width = 40
            
            # Enable text wrapping for all cells
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        logger.info(f"Excel file saved: {self.output_file}")
        logger.info(f"Total emails extracted: {len(self.emails_data)}")
        logger.info(f"Maximum replies in thread: {max_replies}")
    
    def run(self):
        """Main execution method"""
        logger.info("=" * 80)
        logger.info("Starting Epstein Email Extraction Process")
        logger.info("=" * 80)
        logger.info(f"Base directory: {self.base_dir}")
        logger.info(f"Output file: {self.output_file}")
        
        start_time = datetime.now()
        
        try:
            self.process_all_pdfs()
            self.save_to_excel()
            
            end_time = datetime.now()
            duration = end_time - start_time
            
            logger.info("=" * 80)
            logger.info("Extraction Complete!")
            logger.info(f"Total time: {duration}")
            logger.info(f"Total emails extracted: {len(self.emails_data)}")
            logger.info(f"Output file: {self.output_file}")
            logger.info("=" * 80)
            
        except Exception as e:
            logger.error(f"Fatal error during extraction: {e}")
            logger.error(traceback.format_exc())
            raise


def main():
    """Main entry point"""
    base_dir = r"G:\My Drive\04_Resources\Notes\Epstein Email Dump"
    
    extractor = EmailExtractor(base_dir)
    extractor.run()


if __name__ == "__main__":
    main()
